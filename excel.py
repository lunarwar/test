import xlrd
import random
import string
import datetime
import cx_Oracle
import os
from datetime import date, timedelta
from distutils.dir_util import copy_tree
import time as sl
import openpyxl


def open_raw_incident():
    path = "C:/Users/ew/Downloads/"
    arr = os.listdir(path)
    for file in arr:
        if file[:15] == "Problem Ticket_":
            dd = path + file
            raw_data = xlrd.open_workbook(path + file)
            sheet = raw_data.sheet_by_name('PHC')
            return sheet

def insertDB(x,colname,coldif):
    sheet = open_raw_incident()

    wbb = openpyxl.load_workbook("C:/Users/ew6/Desktop/Template.xlsx")
    sheetz = wbb['Templates']
    for r in range(1, 2194):
        column_lenght = sheet.cell(r, 10).value

        if sheet.cell(r, x).value != "":
            OrderId = sheet.cell(r, x).value
            if column_lenght >= coldif:
                sheetz[colname + str(r)] = OrderId

        if sheet.cell(r, x).value == "":
            if column_lenght >= coldif:
                sheetz[colname + str(r)] = OrderId
            print(OrderId)

        wbb.save('C:/Users/ewx510986/Desktop/Template.xlsx')

readCol = [0,1,2,3,4,5,6,7,8,9]
newColName = ['A','B','C','D','E','F','G','H','I','J']
colDiff = [1,1,2,3,4,5,6,7,8]

for i in range(11):
    insertDB(readCol[i],newColName[i],colDiff[i])

